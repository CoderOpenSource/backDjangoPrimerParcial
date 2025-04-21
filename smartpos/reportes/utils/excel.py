import io
import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from productos.models import Producto, MovimientoInventario
from ventas.models import DetalleVenta
from collections import Counter


from io import BytesIO
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from django.contrib.auth import get_user_model

def generar_excel_ventas_por_fecha(ventas, fecha_inicio=None, fecha_fin=None, columnas_extra=None, usuario=None):
    User = get_user_model()
    columnas_extra = columnas_extra or []
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="800080")  # púrpura
    cell_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Rango y cliente
    rango = "Todas las ventas"
    if fecha_inicio and fecha_fin:
        rango = f"Del {fecha_inicio} al {fecha_fin}"
    elif fecha_inicio:
        rango = f"Desde el {fecha_inicio}"
    elif fecha_fin:
        rango = f"Hasta el {fecha_fin}"

    cliente_str = f"{usuario.first_name} {usuario.last_name}" if usuario else "Todos los clientes"
    ws.append([f"Rango de fechas: {rango}", f"Cliente: {cliente_str}"])
    ws.append([])

    for venta in ventas:
        cliente = venta.usuario if hasattr(venta, 'usuario') else None
        nombre = f"{cliente.first_name} {cliente.last_name}" if cliente else '---'
        ci = cliente.ci if cliente and hasattr(cliente, 'ci') else '---'
        celular = cliente.celular if cliente and hasattr(cliente, 'celular') else '---'
        direccion = cliente.direccion if cliente and hasattr(cliente, 'direccion') else '---'
        metodo = venta.tipo_pago.nombre if venta.tipo_pago else '---'
        estado = venta.estado if hasattr(venta, 'estado') else '---'
        fecha = venta.fecha_venta.strftime('%d/%m/%Y %H:%M') if hasattr(venta, 'fecha_venta') and venta.fecha_venta else '---'

        # Tabla resumen
        resumen = [
            ['Venta ID', venta.id, 'Cliente', nombre],
            ['CI Cliente', ci, 'Celular', celular],
            ['Dirección', direccion, 'Pago', metodo],
            ['Estado', estado, 'Fecha', fecha],
            ['Total', f"{venta.total:.2f} Bs", '', '']
        ]
        for row in resumen:
            ws.append(row)

        # Formato tabla resumen
        for row in ws.iter_rows(min_row=ws.max_row - len(resumen) + 1, max_row=ws.max_row):
            for i, cell in enumerate(row):
                cell.border = cell_border
                if i in [0, 2]:  # columnas de etiquetas
                    cell.fill = header_fill
                    cell.font = header_font

        ws.append([])

        # Productos
        ws.append(["Productos asociados a la venta:"])
        header = ["Producto", "Cantidad", "Precio Unitario", "Subtotal"]
        ws.append(header)

        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="center")

        for detalle in venta.detalles.all():
            row = [
                detalle.producto.nombre,
                detalle.cantidad,
                f"{detalle.precio_unitario:.2f} Bs",
                f"{detalle.precio_unitario * detalle.cantidad:.2f} Bs"
            ]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.border = cell_border
                cell.alignment = Alignment(horizontal="center")

        # Total al final
        ws.append(['', '', 'Total:', f"{venta.total:.2f} Bs"])
        for cell in ws[ws.max_row]:
            cell.border = cell_border
            if cell.column in [3, 4]:
                cell.font = Font(bold=True)
        ws.append([])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output




def generar_excel_productos_mas_vendidos():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos más vendidos"

    ws.append(['Producto', 'Cantidad total vendida'])

    detalles = DetalleVenta.objects.select_related('producto')
    contador = Counter()
    for d in detalles:
        contador[d.producto.nombre] += d.cantidad

    for nombre, cantidad in contador.most_common():
        ws.append([nombre, cantidad])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=productos_mas_vendidos.xlsx'
    return response


def generar_excel_inventario():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    ws.append(['Producto', 'Categoría', 'Subcategoría', 'Stock', 'Punto Reorden'])

    productos = Producto.objects.select_related('stock', 'subcategoria__categoria')

    for p in productos:
        stock = getattr(p, 'stock', None)
        ws.append([
            p.nombre,
            p.subcategoria.categoria.nombre if p.subcategoria and p.subcategoria.categoria else '—',
            p.subcategoria.nombre if p.subcategoria else '—',
            stock.cantidad_actual if stock else 0,
            stock.punto_reorden if stock else '—'
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=reporte_inventario.xlsx'
    return response

from io import BytesIO
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse

def generar_excel_productos_mas_vendidos(productos, ventas_dict, fecha_inicio=None, fecha_fin=None,
                                         categoria_nombre=None, subcategoria_nombre=None, marca_nombre=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos más vendidos"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="800080")
    cell_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Información de filtros
    rango = "Todas las fechas"
    if fecha_inicio and fecha_fin:
        rango = f"Del {fecha_inicio} al {fecha_fin}"
    elif fecha_inicio:
        rango = f"Desde el {fecha_inicio}"
    elif fecha_fin:
        rango = f"Hasta el {fecha_fin}"

    cat = categoria_nombre or "Todas"
    subcat = subcategoria_nombre or "Todas"
    marca = marca_nombre or "Todas"

    ws.append([f"Rango de fechas: {rango}"])
    ws.append([f"Categoría: {cat}", f"Subcategoría: {subcat}", f"Marca: {marca}"])
    ws.append([])

    # Cabecera
    headers = ['Producto', 'Categoría', 'Subcategoría', 'Marca',
               'Total', 'Ingreso', 'Ganancia', 'Última Venta']
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = cell_border
        cell.alignment = Alignment(horizontal="center")

    total_ingreso = 0
    total_ganancia = 0
    fechas_venta = []
    ventas_por_categoria = defaultdict(int)
    producto_mas_vendido = ('---', 0)

    for producto in productos:
        venta = ventas_dict.get(producto.id, {})
        cantidad = venta.get('total', 0)
        fecha_venta = venta.get('fecha')

        ingreso = float(cantidad) * float(producto.precio_unitario)
        ganancia = float(cantidad) * (float(producto.precio_unitario) - float(producto.precio_compra or 0))

        total_ingreso += ingreso
        total_ganancia += ganancia

        if fecha_venta:
            fechas_venta.append(fecha_venta)

        cat_nombre = producto.subcategoria.categoria.nombre if producto.subcategoria and producto.subcategoria.categoria else '---'
        ventas_por_categoria[cat_nombre] += cantidad

        if cantidad > producto_mas_vendido[1]:
            producto_mas_vendido = (producto.nombre, cantidad)

        row = [
            producto.nombre,
            cat_nombre,
            producto.subcategoria.nombre if producto.subcategoria else '---',
            producto.marca.nombre if producto.marca else '---',
            cantidad,
            f"{ingreso:.2f} Bs",
            f"{ganancia:.2f} Bs",
            fecha_venta.strftime('%d/%m/%Y %H:%M') if fecha_venta else '---'
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="center")

    # Resumen
    ws.append([])
    ws.append(["Resumen General"])
    if fecha_inicio and fecha_fin:
        resumen = f"Total de ganancia generada en las fechas: {fecha_inicio} al {fecha_fin}"
    elif fechas_venta:
        fechas_venta.sort()
        resumen = f"Total de ganancia generada desde: {fechas_venta[0].strftime('%d/%m/%Y')} hasta {fechas_venta[-1].strftime('%d/%m/%Y')}"
    else:
        resumen = "Total de ganancia generada: sin fechas registradas"
    ws.append([resumen])
    ws.append([f"Total Ingreso: {total_ingreso:.2f} Bs"])
    ws.append([f"Total Ganancia: {total_ganancia:.2f} Bs"])
    ws.append([f"Categoría más vendida: {max(ventas_por_categoria.items(), key=lambda x: x[1])[0] if ventas_por_categoria else '---'}"])
    ws.append([f"Producto más vendido: {producto_mas_vendido[0]}"])

    # Ajuste ancho columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict


def generar_excel_inventario(movimientos):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Inventario"
    ws.title = "Reporte de Inventario"

    if not movimientos:
        ws.append(["No se encontraron movimientos con los filtros aplicados."])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="800080")
    cell_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Encabezado
    headers = ['Producto', 'Categoría', 'Subcategoría', 'Marca', 'Tipo', 'Cantidad', 'Fecha', 'Descripción']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = cell_border

    totales = {'entrada': 0, 'salida': 0}
    productos_afectados = set()
    fechas = []
    conteo_por_categoria = defaultdict(int)

    for m in movimientos:
        producto = m.producto
        productos_afectados.add(producto.id)
        cat = producto.subcategoria.categoria.nombre if producto.subcategoria and producto.subcategoria.categoria else '---'
        subcat = producto.subcategoria.nombre if producto.subcategoria else '---'
        marca = producto.marca.nombre if producto.marca else '---'
        fecha = m.fecha_movimiento.strftime('%d/%m/%Y %H:%M') if m.fecha_movimiento else '---'

        row = [
            producto.nombre,
            cat,
            subcat,
            marca,
            m.tipo.capitalize(),
            m.cantidad,
            fecha,
            m.descripcion or '---'
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="center")

        totales[m.tipo] += m.cantidad
        conteo_por_categoria[cat] += m.cantidad
        fechas.append(m.fecha_movimiento)

    # Espacio vacío
    ws.append([])

    # Resumen
    fechas.sort()
    if fechas:
        resumen = f"Resumen de movimientos desde: {fechas[0].strftime('%d/%m/%Y')} hasta {fechas[-1].strftime('%d/%m/%Y')}"
    else:
        resumen = "Resumen de movimientos: sin fechas registradas"

    categoria_top = max(conteo_por_categoria.items(), key=lambda x: x[1])[0] if conteo_por_categoria else '---'

    resumen_data = [
        [resumen],
        [f"Total de productos distintos con movimientos: {len(productos_afectados)}"],
        [f"Total Entradas: {totales['entrada']} unidades"],
        [f"Total Salidas: {totales['salida']} unidades"],
        [f"Categoría con más movimientos: {categoria_top}"]
    ]
    for row in resumen_data:
        ws.append(row)

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output